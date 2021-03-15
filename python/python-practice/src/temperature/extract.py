import csv
from pathlib import Path
from openpyxl import load_workbook
from transform import transfer_words
from clean import clean_line
import save

header_list=[]
content_list={}

def get_line(path):
    file=Path(path)
    with open(file,'r') as content:
        content_list=create_content_list(content.readlines())
    return content_list

def create_content_list(lines):
    global content_list,header_list
    line_index=0
    for line in lines:
        words=transfer_words(clean_line(line))
        if line_index ==0:
            header_list=words
        else:
            content_list.setdefault(line_index)
            content_list[line_index]=words
        line_index+=1
    return content_list

def getdata(path):
    file_path=Path(path)
    with open(file_path,'r') as file:
        results=[fields for fields in csv.reader(file,delimiter="|")]
    return results

def readcsv(path):
    csv_path=Path(path)
    with open(csv_path,newline='',encoding="utf-8") as content:
        results=[fields for fields in csv.reader(content)]
    return results

def readxls(path):
    wb=load_workbook(path)
    results=[]
    body_tempure_list={}
    ws=wb.worksheets[0]
    row_index=0
    for row in ws.iter_rows():
        # cell.value if cell.value !="Missing" else 0 for cell in row
        results.append([ cell.value for cell in row])
        if row_index>0:
            body_tempure_list.setdefault(row[4].value)
            body_tempure_list[row[4].value]=[float(cell.value) if cell.value!="Missing" else "Missing" for cell in row[11:13] ]
        row_index+=1
    print(body_tempure_list)

# readxls("../../sample/temp_data_01.xlsx")

def main():
    contents=get_line("../../sample/temp_data_pipes_00a.txt")
    table_name="temperature"
    create_sytax="""create table {0} (
                         id integer primary key,
                         city text,
                         date text,
                         temperature float,
                         count integer)""".format(table_name)
    save.create_table(table_name,create_sytax)
    if len(save.select(table_name))==0:
        row_index=1
        for data in  contents.values():
            save.insert( table_name,row_index,data[0], data[1], data[2],data[3],"id","city","date","temperature","count")
            row_index+=1
    print(save.select(table_name))

if __name__=="__main__":
    main()