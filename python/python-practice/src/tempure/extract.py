import csv
from pathlib import Path
from unittest import result
from openpyxl import load_workbook

def getdata(path):
    file_path=Path(path)
    with open(file_path,'r') as file:
        results=[fields for fields in csv.reader(file,delimiter="|")]
    return results

def readcsv(path):
    csv_path=Path(path)
    with open(csv_path,newline='',encoding="utf-8") as content:
        results=[fields for fields in csv.reader(content)]
    print(results) 

def readxls(path):
    wb=load_workbook(path)
    results=[]
    body_tempure_list={}
    ws=wb.worksheets[0]
    row_index=0
    for row in ws.iter_rows():
        # cell.value if cell.value !="Missing" else 0 for cell in row
        results.append([ cell.value for cell in row])
        if row_index>0:
            body_tempure_list.setdefault(row[4].value)
            body_tempure_list[row[4].value]=[float(cell.value) if cell.value!="Missing" else "Missing" for cell in row[11:13] ]
        row_index+=1
    print(body_tempure_list)

readxls("../../sample/temp_data_01.xlsx")